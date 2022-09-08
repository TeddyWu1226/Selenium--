#!/usr/bin/python
# -*- coding: UTF-8 -*-
"""
Author:Cheng Hong,Wu
專案:查找證券交易所資料
資料來源:https://www.twse.com.tw/zh/page/trading/exchange/MI_INDEX.html
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import Select
import openpyxl
import time
import datetime

import warnings
warnings.filterwarnings("ignore")

cc = 0
print("台灣證券交易所-股票資料下載\n資料來源:https://www.twse.com.tw/zh/page/trading/exchange/MI_INDEX.html\n")
print("類別代碼編號表:\n0大盤統計資訊、1收盤指數資訊、2委託及成交統計資訊、3全部、4全部(不含權證、牛熊證、可展延牛熊證)、5封閉式基金、6ETF、7ETN、8受益證券、9認購權證(不含牛證)\n10認售權證(不含熊證)、11牛證(不含可展延牛證)、12熊證(不含可展延熊證)、13可展延牛證、14可展延熊證、15附認股權特別股、16附認股權公司債、17認股權憑證、18可轉換公司債、19創新板股票\n20水泥工業、21食品工業、22塑膠工業、23紡織纖維、24電機機械、25電器電纜、26化學生技醫療、27化學工業、28生技醫療業、29玻璃陶瓷\n30造紙工業、31鋼鐵工業、32橡膠工業、33汽車工業、34電子工業、35半導體業、36電腦及週邊設備業、37光電業、38通信網路業、39電子零組件業\n40電子通路業、41資訊服務業、42其他電子業、43建材營造、44航運業、45觀光事業、46金融保險、47貿易百貨、48存託憑證、49油電燃氣業\n50綜合、51其他\n")
while True:
    try:
        cc = int(input("請輸入類別代碼(ex:5):"))
        break
    except Exception:
        print("格式輸入錯誤，請重新輸入")
while True:
    try:
        yy = int(input("請輸入年(2004~西元今年):"))
        break
    except Exception:
        print("格式輸入錯誤，請重新輸入")
while True:
    try:
        mm = int(input("請輸入月(1~12):"))
        break
    except Exception:
        print("格式輸入錯誤，請重新輸入")
while True:
    try:
        dd = int(input("請輸入日(0~31):"))
        break
    except Exception:
        print("格式輸入錯誤，請重新輸入")
yy = int(datetime.date.today().strftime('%Y')) - yy
# setting
def selenium_selection(category=35, year=0, month=0, date=3):  # 從零開始數
    # 啟動 selenium chrome 開啟網頁
    option = webdriver.ChromeOptions()  # 選擇chrome開啟
    driver = webdriver.Chrome('chromedriver.exe', options=option)
    driver.get('https://www.twse.com.tw/zh/page/trading/exchange/MI_INDEX.html')
    time.sleep(0.1)
    # 選擇產業()
    try:
        select = Select(driver.find_element(By.NAME, 'type'))  # 類型分類為type
        select.select_by_index(category)
    except Exception:
        print("爬蟲遭阻擋或該網頁已無效，請等一下再嘗試QQ")
        driver.close()  # 關掉網頁
        exit()
    # 選擇年分 (0~18),18=民國93年(2004)
    select = Select(driver.find_element(By.NAME, 'yy'))  # 類型分類為yy
    select.select_by_index(year)
    # 選擇月份(0~11)
    select = Select(driver.find_element(By.NAME, 'mm'))  # 產業類型分類為mm
    select.select_by_index(month)
    # 選擇日期(0~27,29or30)
    select = Select(driver.find_element(By.NAME, 'dd'))  # 產業類型分類為dd
    select.select_by_index(date)
    # 按下搜尋
    element = driver.find_elements(By.CLASS_NAME, "search")  # button search找不到
    element[1].click()
    time.sleep(0.2)  # 給他跑一下
    # 判斷是否有資料

    soup = BeautifulSoup(driver.page_source, 'html.parser')
    #print(soup)
    soup.find(id="result-message")
    try:
        test = soup.select("#result-message")[0].string
    except AttributeError:
        print("無資料，請重新選擇日期")
        driver.close()  # 關掉網頁
        exit()
    # 展開所有股票
    select = Select(driver.find_element(By.NAME, 'report-table1_length'))  # 類型分類為report-table1_length
    select.select_by_index(4)

    # BS4
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    # BS4讀取資料 ,list 化
    colName = []
    dataList = []

    title = soup.select("#subtitle1")[0].string
    data = soup.select(".data-table")[0].find_all("th")
    for i in range(len(data)):
        colName.append(data[i].string)
    data = soup.select(".data-table")[0].find_all("td")
    for i in range(len(data)):
        dataList.append(data[i].string)
    #driver.close()   # 關掉網頁


    if '(元,股)' in colName:
        colName.remove('(元,股)')
    if '(元,交易單位)' in colName:
        colName.remove('(元,交易單位)')

    # 資料寫入xlsx
    wb = openpyxl.Workbook()
    sheetMax = 0
    selectSheet = wb.create_sheet(f"{int(datetime.date.today().strftime('%Y'))-year}.{1+month}.{1+date}", sheetMax)
    # 放標題
    for i in range(len(colName)):
        selectSheet.cell(row=1, column=i+1, value=colName[i])

    # 放資料
    n = 2
    colLimit = 0
    for i in dataList:
        selectSheet.cell(row=n, column=colLimit+1, value=i)
        colLimit += 1
        if colLimit == len(colName):
            n += 1
            colLimit = 0
        else:
            pass

    wb.save(f"台灣證券交易所-{title}.xlsx")
    print(f"台灣證券交易所-{title}.xlsx檔案已建立")


selenium_selection(cc,yy,mm-1,dd-1)
